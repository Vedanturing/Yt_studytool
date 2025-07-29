import os
import json
import subprocess
import logging
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel
from googleapiclient.discovery import build
from dotenv import load_dotenv
import io
import tempfile
import asyncio
import urllib.parse
from pathlib import Path
import re
import sqlite3
import hashlib

# Import export libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    EXPORT_AVAILABLE = True
except ImportError:
    EXPORT_AVAILABLE = False
    print("Warning: Export libraries not available. Install openpyxl and reportlab for export functionality.")

# Import transcription and AI libraries
try:
    import whisper
    import openai
    from transformers import pipeline
    import google.generativeai as genai
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    print("Warning: AI libraries not available. Install whisper, openai, transformers, and google-generativeai for AI functionality.")

# Import syllabus parsing libraries
try:
    import PyPDF2
    import docx
    SYLLABUS_PARSING_AVAILABLE = True
except ImportError:
    SYLLABUS_PARSING_AVAILABLE = False
    print("Warning: Syllabus parsing libraries not available. Install PyPDF2 and python-docx for syllabus functionality.")

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import study routes with robust import handling
STUDY_ROUTES_AVAILABLE = False
study_router = None

try:
    # Try relative import first (when running as module)
    from .study_routes import router as study_router
    STUDY_ROUTES_AVAILABLE = True
    print("✅ Study routes imported successfully (relative import)")
except ImportError:
    try:
        # Try absolute import (when running directly)
        from study_routes import router as study_router
        STUDY_ROUTES_AVAILABLE = True
        print("✅ Study routes imported successfully (absolute import)")
    except ImportError:
        try:
            # Try with backend prefix (when running from project root)
            from backend.study_routes import router as study_router
            STUDY_ROUTES_AVAILABLE = True
            print("✅ Study routes imported successfully (backend prefix)")
        except ImportError as e:
            STUDY_ROUTES_AVAILABLE = False
            print(f"⚠️  Study routes not available: {e}")
            print("   The study module will not be available.")

app = FastAPI(title="YouTube Video Search API", version="1.0.0")

# Add CORS middleware
try:
    from cors_config import get_fastapi_cors_config
    cors_config = get_fastapi_cors_config()
    app.add_middleware(CORSMiddleware, **cors_config)
    print("✅ FastAPI CORS configured with flexible origins")
except ImportError:
    # Fallback to basic CORS configuration
    app.add_middleware(
        CORSMiddleware,
        allow_origins=[
            "http://localhost:3000", 
            "http://127.0.0.1:3000",
            "http://localhost:3001", 
            "http://127.0.0.1:3001"
        ],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    print("⚠️  FastAPI using fallback CORS configuration")

# Include study routes if available
if STUDY_ROUTES_AVAILABLE and study_router is not None:
    app.include_router(study_router)
    print("✅ Study routes included in FastAPI app")
else:
    print("⚠️  Study routes not included - module not available")

# Global variables for AI models
whisper_model = None
summarizer = None

# Core Pydantic models
class Comment(BaseModel):
    text: str
    author: str
    likes: int = 0

class Video(BaseModel):
    title: str
    video_url: str
    views: int
    likes: int
    description: str
    comment_count: int
    top_comments: List[Comment]
    thumbnail_url: str = ""
    duration: Optional[float] = None  # Duration in seconds

# New Pydantic models for syllabus feature
class SyllabusTopic(BaseModel):
    unit: str
    topic: str
    description: Optional[str] = None

class SyllabusUploadRequest(BaseModel):
    topics: List[SyllabusTopic]

class SyllabusVideoMapping(BaseModel):
    topic: str
    unit: str
    videos: List[Video]

class SyllabusVideosResponse(BaseModel):
    syllabus_mapping: List[SyllabusVideoMapping]
    total_topics: int
    total_videos: int

class QuizQuestion(BaseModel):
    question: str
    options: List[str]
    answer: str
    explanation: Optional[str] = None
    topic: str
    difficulty: str = "medium"  # "easy", "medium", "hard"
    type: str = "mcq"  # "mcq", "true_false", "fill_blank", "code_output"

class QuizRequest(BaseModel):
    topics: List[str]
    num_questions: int = 20
    question_types: List[str] = ["mcq", "true_false"]
    difficulty: str = "medium"  # "easy", "medium", "hard"
    subject: str = "General"  # For organizing quizzes by subject

class QuizResponse(BaseModel):
    questions: List[QuizQuestion]
    total_questions: int
    topics_covered: List[str]
    source: str = "api"  # "api", "offline", "fallback"
    subject: str = "General"
    difficulty: str = "medium"
    question_types: List[str] = []

class QuizAttempt(BaseModel):
    question: str
    selected_answer: str
    correct_answer: str
    is_correct: bool
    topic: str

class ReportRequest(BaseModel):
    quiz_attempts: List[QuizAttempt]
    watched_videos: List[str]
    syllabus_topics: List[SyllabusTopic]

class ReportResponse(BaseModel):
    overall_score: float
    topic_scores: dict
    weak_areas: List[str]
    recommendations: List[str]
    common_mistakes: List[str]
    unwatched_topics: List[str]
    report_data: dict

# Existing Pydantic models (reordered)
class VideoRequest(BaseModel):
    keyword: str

class TranscribeRequest(BaseModel):
    video_url: str

class TranscribeResponse(BaseModel):
    transcription: str
    video_url: str
    video_title: str = ""
    duration: Optional[float] = None

class SummarizeRequest(BaseModel):
    transcription: str

class SummarizeResponse(BaseModel):
    summary: str
    original_length: int
    summary_length: int

class Flashcard(BaseModel):
    question: str
    answer: str
    type: str  # 'mcq' or 'fact'

class LearningModeResponse(BaseModel):
    video_id: str
    video_title: str
    flashcards: List[Flashcard]
    total_cards: int

class VideoResponse(BaseModel):
    videos: List[Video]
    total_count: int
    source: str  # "api" or "yt-dlp"
    keyword: str

class ExportRequest(BaseModel):
    videos: List[Video]
    keyword: Optional[str] = ""

class ErrorResponse(BaseModel):
    error: str
    details: Optional[str] = None

def initialize_ai_models():
    """Initialize AI models (Whisper and T5)"""
    global AI_AVAILABLE, whisper_model, summarizer
    
    try:
        logger.info("Initializing AI models...")
        
        # Test Whisper import
        import whisper
        logger.info("Whisper imported successfully")
        
        # Test transformers import
        from transformers import pipeline
        logger.info("Transformers imported successfully")
        
        # Initialize models as None (lazy loading)
        whisper_model = None
        summarizer = None
        AI_AVAILABLE = True
        
        logger.info("AI models loaded successfully")
    except Exception as e:
        logger.error(f"Failed to load AI models: {e}")
        AI_AVAILABLE = False

# Initialize models on startup
# if AI_AVAILABLE:
#     initialize_ai_models()

def get_whisper_model():
    """Lazy load Whisper model"""
    global whisper_model
    if whisper_model is None and AI_AVAILABLE:
        try:
            # Get model size from environment variable, default to 'base'
            model_size = os.getenv("WHISPER_MODEL_SIZE", "base")
            logger.info(f"Loading Whisper model ({model_size})...")
            whisper_model = whisper.load_model(model_size)
            logger.info("Whisper model loaded successfully")
        except Exception as e:
            logger.error(f"Failed to load Whisper model: {e}")
            raise HTTPException(status_code=500, detail="Failed to load Whisper model")
    return whisper_model

def get_summarizer():
    """Lazy load summarization model"""
    global summarizer
    if summarizer is None and AI_AVAILABLE:
        try:
            # Get model from environment variable, default to 't5-small'
            model_name = os.getenv("SUMMARIZATION_MODEL", "t5-small")
            logger.info(f"Loading summarization model ({model_name})...")
            
            # Try Gemini first as primary
            if model_name.lower() in ["gemini", "openai"]:
                gemini_api_key = os.getenv("GEMINI_API_KEY")
                if gemini_api_key:
                    logger.info("Using Gemini API as primary summarization model")
                    return "gemini"
                
                # Try OpenAI as fallback
                openai_api_key = os.getenv("OPENAI_API_KEY")
                if openai_api_key:
                    logger.info("Using OpenAI API as fallback")
                    return "openai"
                
                # If neither API key is available
                logger.warning("Both Gemini and OpenAI API keys not found, falling back to t5-small")
                model_name = "t5-small"
            
            summarizer = pipeline("summarization", model=model_name, device=-1)
            logger.info("Summarization model loaded successfully")
        except Exception as e:
            logger.error(f"Failed to load summarization model: {e}")
            raise HTTPException(status_code=500, detail="Failed to load summarization model")
    return summarizer

def generate_filename(prefix: str, extension: str, keyword: str = "") -> str:
    """Generate filename with timestamp"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = "".join(c for c in keyword if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_keyword = safe_keyword.replace(' ', '_')[:20]  # Limit length
    return f"{prefix}_{safe_keyword}_{timestamp}.{extension}"

def create_excel_file(videos: List[Video], keyword: str = "") -> bytes:
    """Create Excel file with video data"""
    if not EXPORT_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available. Install openpyxl.")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Videos"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Title", "Channel", "Views", "Likes", "Comments", "URL", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, video in enumerate(videos, 2):
        ws.cell(row=row, column=1, value=video.title)
        ws.cell(row=row, column=2, value="YouTube")  # Default channel name
        ws.cell(row=row, column=3, value=video.views)
        ws.cell(row=row, column=4, value=video.likes)
        ws.cell(row=row, column=5, value=video.comment_count)
        ws.cell(row=row, column=6, value=video.video_url)
        ws.cell(row=row, column=7, value=video.description[:500])  # Limit description length
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_pdf_file(videos: List[Video], keyword: str = "") -> bytes:
    """Create PDF file with video data"""
    if not EXPORT_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF export not available. Install reportlab.")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title_text = f"YouTube Video Search Results"
    if keyword:
        title_text += f" - {keyword}"
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 20))
    
    # Table data
    table_data = [["Title", "Views", "Likes", "Comments", "URL"]]
    
    for video in videos:
        # Truncate title if too long
        title = video.title[:50] + "..." if len(video.title) > 50 else video.title
        url = video.video_url[:30] + "..." if len(video.video_url) > 30 else video.video_url
        
        table_data.append([
            title,
            str(video.views),
            str(video.likes),
            str(video.comment_count),
            url
        ])
    
    # Create table
    table = Table(table_data, colWidths=[2.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch])
    
    # Table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ])
    table.setStyle(table_style)
    
    story.append(table)
    doc.build(story)
    
    output.seek(0)
    return output.getvalue()

def extract_video_id(video_url: str) -> str:
    """Extract video ID from YouTube URL"""
    if "youtube.com/watch?v=" in video_url:
        return video_url.split("v=")[1].split("&")[0]
    elif "youtu.be/" in video_url:
        return video_url.split("youtu.be/")[1].split("?")[0]
    else:
        raise ValueError("Invalid YouTube URL format")

def download_audio(video_url: str, output_dir: str) -> tuple[str, str, float]:
    """Download audio from YouTube video using yt-dlp"""
    try:
        video_id = extract_video_id(video_url)
        
        title = "Unknown Title"
        duration = 0.0
        
        # Use the exact same approach that worked in our test
        logger.info("Downloading audio using working approach...")
        output_template = os.path.join(output_dir, f"{video_id}.%(ext)s")
        
        cmd = [
            "yt-dlp",
            "--format", "bestaudio",
            "--output", output_template,
            "--no-playlist",
            "--quiet",
            video_url
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        
        logger.info(f"Download - Return code: {result.returncode}")
        if result.stdout:
            logger.info(f"Download - stdout: {result.stdout}")
        if result.stderr:
            logger.info(f"Download - stderr: {result.stderr}")
        
        # Get metadata separately to avoid mixing with download output
        logger.info("Getting video metadata...")
        cmd_meta = [
            "yt-dlp",
            "--simulate",
            "--print", "%(title)s",
            "--print", "%(duration)s",
            video_url
        ]
        
        meta_result = subprocess.run(cmd_meta, capture_output=True, text=True, timeout=60)
        
        if meta_result.returncode == 0 and meta_result.stdout.strip():
            meta_lines = meta_result.stdout.strip().split('\n')
            if len(meta_lines) >= 1:
                title = meta_lines[0] or "Unknown Title"
            if len(meta_lines) >= 2:
                try:
                    duration = float(meta_lines[1]) if meta_lines[1] and meta_lines[1].replace('.', '').isdigit() else 0.0
                except:
                    duration = 0.0
            
            logger.info(f"Metadata - Title: {title}, Duration: {duration}")
        
        # If first attempt failed, try alternatives
        if result.returncode != 0:
            logger.info("First download failed, trying alternative formats...")
            
            # Try worst quality
            cmd_worst = [
                "yt-dlp",
                "--format", "worst",
                "--output", output_template,
                "--no-playlist",
                "--quiet",
                video_url
            ]
            
            result = subprocess.run(cmd_worst, capture_output=True, text=True, timeout=300)
            
            if result.returncode != 0:
                # Try any format
                cmd_any = [
                    "yt-dlp",
                    "--output", output_template,
                    "--no-playlist",
                    "--quiet",
                    video_url
                ]
                
                result = subprocess.run(cmd_any, capture_output=True, text=True, timeout=300)
        
        # Find the downloaded file
        audio_file = None
        
        try:
            files = os.listdir(output_dir)
            logger.info(f"Files in directory after download: {files}")
            
            # Look for files with the video ID first
            for file in files:
                if video_id in file:
                    audio_file = os.path.join(output_dir, file)
                    logger.info(f"Found file with video ID: {file}")
                    break
            
            # If no file with video ID, look for any media file
            if not audio_file:
                for file in files:
                    if any(file.lower().endswith(ext) for ext in ['.webm', '.m4a', '.mp3', '.wav', '.ogg', '.mp4', '.mkv', '.aac', '.flv']):
                        audio_file = os.path.join(output_dir, file)
                        logger.info(f"Found media file: {file}")
                        break
            
            # Last resort: take any file in the directory
            if not audio_file and files:
                audio_file = os.path.join(output_dir, files[0])
                logger.info(f"Using first available file: {files[0]}")
                
        except Exception as e:
            logger.error(f"Error listing directory: {e}")
        
        if not audio_file or not os.path.exists(audio_file):
            # Final attempt with verbose output to debug
            logger.error("No files found. Attempting download with verbose output...")
            cmd_verbose = [
                "yt-dlp",
                "--format", "bestaudio",
                "--output", output_template,
                "--no-playlist",
                "-v",  # Verbose output
                video_url
            ]
            
            verbose_result = subprocess.run(cmd_verbose, capture_output=True, text=True, timeout=300)
            logger.error(f"Verbose download output: {verbose_result.stdout}")
            logger.error(f"Verbose download errors: {verbose_result.stderr}")
            
            raise Exception("No audio file found after download attempts. YouTube may be blocking requests or the video may not be available.")
        
        # Verify file is not empty
        file_size = os.path.getsize(audio_file)
        if file_size == 0:
            raise Exception("Downloaded file is empty")
        
        logger.info(f"Successfully found audio file: {audio_file} ({file_size} bytes)")
        logger.info(f"Video title: {title}, Duration: {duration}")
        
        return audio_file, title, duration
        
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=500, detail="Download timeout. The video might be too long or the connection is slow.")
    except Exception as e:
        logger.error(f"Error downloading audio: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download audio: {str(e)}")

def transcribe_with_gemini(audio_file: str) -> str:
    """Transcribe audio using Gemini API (experimental)"""
    try:
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if not gemini_api_key:
            raise Exception("Gemini API key not found")
        
        # Configure Gemini
        genai.configure(api_key=gemini_api_key)
        
        # Note: Gemini doesn't directly support audio transcription yet
        # This is a placeholder for future functionality
        # For now, we'll fall back to Whisper
        raise Exception("Gemini audio transcription not yet supported")
        
    except Exception as e:
        logger.warning(f"Gemini transcription failed: {e}")
        raise e

def transcribe_audio(audio_file: str) -> str:
    """Transcribe audio file using Whisper (primary) with Gemini fallback planned"""
    if not AI_AVAILABLE:
        raise HTTPException(status_code=500, detail="AI libraries not available. Please install whisper.")
    
    try:
        logger.info(f"Transcribing audio file: {audio_file}")
        
        # Check if file exists
        if not os.path.exists(audio_file):
            raise Exception(f"Audio file not found: {audio_file}")
        
        # Check file size
        file_size = os.path.getsize(audio_file)
        if file_size == 0:
            raise Exception("Audio file is empty")
        
        # Check if file is too large (> 100MB)
        max_size = 100 * 1024 * 1024  # 100MB
        if file_size > max_size:
            logger.warning(f"Audio file is large ({file_size / 1024 / 1024:.1f}MB). This may take a while.")
        
        logger.info(f"Audio file size: {file_size} bytes")
        
        # Use Whisper as primary transcription method
        model = get_whisper_model()
        logger.info("Starting Whisper transcription...")
        
        # Add better error handling for transcription
        try:
            result = model.transcribe(
                audio_file,
                fp16=False,  # Use fp32 for better compatibility
                temperature=0.0,  # More deterministic output
                beam_size=5,  # Better quality
                best_of=5,  # More attempts for best result
                patience=1.0
            )
            transcription = result["text"].strip()
        except Exception as whisper_error:
            logger.error(f"Whisper transcription failed: {whisper_error}")
            # Try with minimal settings as fallback
            result = model.transcribe(audio_file, fp16=False)
            transcription = result["text"].strip()
        
        if not transcription:
            raise Exception("Transcription resulted in empty text. The audio might be silent or corrupted.")
        
        # Clean up the transcription
        transcription = transcription.strip()
        if len(transcription) < 10:
            logger.warning(f"Transcription is very short: '{transcription}'")
        
        logger.info(f"Transcription completed. Length: {len(transcription)} characters")
        return transcription
        
    except Exception as e:
        logger.error(f"Error transcribing audio: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to transcribe audio: {str(e)}")

def chunk_text(text: str, max_length: int = 512) -> List[str]:
    """Split text into chunks for summarization"""
    words = text.split()
    chunks = []
    current_chunk = []
    current_length = 0
    
    for word in words:
        if current_length + len(word) + 1 > max_length and current_chunk:
            chunks.append(' '.join(current_chunk))
            current_chunk = [word]
            current_length = len(word)
        else:
            current_chunk.append(word)
            current_length += len(word) + 1
    
    if current_chunk:
        chunks.append(' '.join(current_chunk))
    
    return chunks

def summarize_text(text: str) -> str:
    """Summarize text using transformer model or OpenAI API"""
    if not AI_AVAILABLE:
        raise HTTPException(status_code=500, detail="Summarization model not available")
    
    try:
        # Handle empty or very short text
        if len(text.strip()) < 50:
            return text.strip()
        
        summarizer_model = get_summarizer()
        
        # Check if using Gemini API (primary)
        if summarizer_model == "gemini":
            try:
                return summarize_with_gemini(text)
            except Exception as e:
                logger.warning(f"Gemini API failed, trying OpenAI: {e}")
                try:
                    return summarize_with_openai(text)
                except Exception as e2:
                    logger.warning(f"OpenAI API also failed: {e2}")
                    # Fall through to local model
        
        # Check if using OpenAI API (fallback)
        if summarizer_model == "openai":
            try:
                return summarize_with_openai(text)
            except Exception as e:
                logger.warning(f"OpenAI API failed, trying Gemini: {e}")
                return summarize_with_gemini(text)
        
        # For very long text, chunk it and summarize each chunk
        if len(text) > 1024:
            chunks = chunk_text(text, 512)
            summaries = []
            
            for chunk in chunks:
                if len(chunk.strip()) > 50:  # Only summarize chunks with substantial content
                    try:
                        result = summarizer_model(chunk, max_length=150, min_length=30, do_sample=False)
                        summaries.append(result[0]['summary_text'])
                    except Exception as e:
                        logger.warning(f"Failed to summarize chunk: {e}")
                        summaries.append(chunk[:200] + "...")  # Fallback to truncation
            
            # Combine summaries
            combined_summary = " ".join(summaries)
            
            # If combined summary is still too long, summarize it again
            if len(combined_summary) > 1024:
                try:
                    final_result = summarizer_model(combined_summary, max_length=300, min_length=100, do_sample=False)
                    return final_result[0]['summary_text']
                except:
                    return combined_summary[:500] + "..."  # Fallback
            
            return combined_summary
        else:
            # Direct summarization for shorter text
            result = summarizer_model(text, max_length=200, min_length=50, do_sample=False)
            return result[0]['summary_text']
            
    except Exception as e:
        logger.error(f"Error summarizing text: {e}")
        # Fallback to simple truncation
        sentences = text.split('. ')
        if len(sentences) > 3:
            return '. '.join(sentences[:3]) + '.'
        return text[:500] + "..." if len(text) > 500 else text

def generate_flashcards_with_gemini(text: str, video_title: str) -> List[Flashcard]:
    """Generate flashcards using Gemini API"""
    try:
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if not gemini_api_key:
            raise Exception("Gemini API key not found")
        
        # Configure Gemini
        genai.configure(api_key=gemini_api_key)
        model = genai.GenerativeModel('gemini-pro')
        
        # Create a comprehensive prompt for flashcard generation
        prompt = f"""
Based on the following video transcription titled "{video_title}", create 3-5 educational flashcards in JSON format.

Generate a mix of factual questions and multiple-choice questions that test key concepts from the content.

Return ONLY a valid JSON array with this exact structure:
[
  {{
    "question": "What is the main topic discussed?",
    "answer": "Brief, clear answer",
    "type": "fact"
  }},
  {{
    "question": "Which of the following best describes X? A) Option 1 B) Option 2 C) Option 3 D) Option 4",
    "answer": "C) Option 3 - because explanation",
    "type": "mcq"
  }}
]

Transcription:
{text[:3000]}...
"""
        
        response = model.generate_content(prompt)
        response_text = response.text.strip()
        
        # Try to extract JSON from the response
        import json
        import re
        
        # Look for JSON array in the response
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group()
            try:
                flashcards_data = json.loads(json_str)
                flashcards = []
                for card_data in flashcards_data:
                    if isinstance(card_data, dict) and 'question' in card_data and 'answer' in card_data:
                        flashcards.append(Flashcard(
                            question=card_data.get('question', ''),
                            answer=card_data.get('answer', ''),
                            type=card_data.get('type', 'fact')
                        ))
                return flashcards
            except json.JSONDecodeError:
                pass
        
        # Fallback: create basic flashcards from text
        return generate_fallback_flashcards(text, video_title)
        
    except Exception as e:
        logger.error(f"Error generating flashcards with Gemini: {e}")
        return generate_fallback_flashcards(text, video_title)

def generate_fallback_flashcards(text: str, video_title: str) -> List[Flashcard]:
    """Generate basic flashcards as fallback"""
    flashcards = []
    
    # Extract key sentences and create basic Q&A
    sentences = [s.strip() for s in text.split('.') if len(s.strip()) > 50][:10]
    
    # Create a few basic flashcards
    flashcards.append(Flashcard(
        question=f"What is the main topic of the video '{video_title}'?",
        answer=f"The video discusses: {sentences[0][:100]}..." if sentences else "Content analysis",
        type="fact"
    ))
    
    if len(sentences) > 2:
        flashcards.append(Flashcard(
            question="What key concept is explained in this video?",
            answer=sentences[1][:150] + "..." if len(sentences[1]) > 150 else sentences[1],
            type="fact"
        ))
    
    if len(sentences) > 4:
        flashcards.append(Flashcard(
            question=f"Which statement best describes the content? A) Unrelated topic B) {sentences[2][:50]}... C) Different subject D) None of the above",
            answer=f"B) {sentences[2][:50]}... - This directly relates to the video content",
            type="mcq"
        ))
    
    return flashcards

def summarize_with_gemini(text: str) -> str:
    """Summarize text using Google Gemini API"""
    try:
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if not gemini_api_key:
            raise Exception("Gemini API key not found")
        
        # Configure Gemini
        genai.configure(api_key=gemini_api_key)
        model = genai.GenerativeModel('gemini-pro')
        
        # For very long text, chunk it first
        if len(text) > 4000:
            chunks = chunk_text(text, 3000)
            summaries = []
            
            for chunk in chunks:
                if len(chunk.strip()) > 100:
                    try:
                        prompt = f"Please summarize the following text in 2-3 sentences:\n\n{chunk}"
                        response = model.generate_content(prompt)
                        summaries.append(response.text.strip())
                    except Exception as e:
                        logger.warning(f"Failed to summarize chunk with Gemini: {e}")
                        summaries.append(chunk[:200] + "...")
            
            combined_summary = " ".join(summaries)
            
            # If still too long, summarize the combined summary
            if len(combined_summary) > 1000:
                prompt = f"Please provide a concise summary of the following text:\n\n{combined_summary}"
                response = model.generate_content(prompt)
                return response.text.strip()
            
            return combined_summary
        else:
            # Direct summarization for shorter text
            prompt = f"Please summarize the following text in 2-3 sentences:\n\n{text}"
            response = model.generate_content(prompt)
            return response.text.strip()
            
    except Exception as e:
        logger.error(f"Error using Gemini API: {e}")
        # Fallback to simple truncation
        sentences = text.split('. ')
        if len(sentences) > 3:
            return '. '.join(sentences[:3]) + '.'
        return text[:500] + "..." if len(text) > 500 else text

def summarize_with_openai(text: str) -> str:
    """Summarize text using OpenAI API"""
    try:
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise Exception("OpenAI API key not found")
        
        # Initialize OpenAI client
        client = openai.OpenAI(api_key=openai_api_key)
        
        # For very long text, chunk it first
        if len(text) > 4000:
            chunks = chunk_text(text, 3000)
            summaries = []
            
            for chunk in chunks:
                if len(chunk.strip()) > 100:
                    try:
                        response = client.chat.completions.create(
                            model="gpt-3.5-turbo",
                            messages=[
                                {"role": "system", "content": "You are a helpful assistant that summarizes text concisely and accurately."},
                                {"role": "user", "content": f"Please summarize the following text in 2-3 sentences:\n\n{chunk}"}
                            ],
                            max_tokens=150,
                            temperature=0.3
                        )
                        summaries.append(response.choices[0].message.content.strip())
                    except Exception as e:
                        logger.warning(f"Failed to summarize chunk with OpenAI: {e}")
                        summaries.append(chunk[:200] + "...")
            
            combined_summary = " ".join(summaries)
            
            # If still too long, summarize the combined summary
            if len(combined_summary) > 1000:
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant that summarizes text concisely and accurately."},
                        {"role": "user", "content": f"Please provide a concise summary of the following text:\n\n{combined_summary}"}
                    ],
                    max_tokens=200,
                    temperature=0.3
                )
                return response.choices[0].message.content.strip()
            
            return combined_summary
        else:
            # Direct summarization for shorter text
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that summarizes text concisely and accurately."},
                    {"role": "user", "content": f"Please summarize the following text in 2-3 sentences:\n\n{text}"}
                ],
                max_tokens=200,
                temperature=0.3
            )
            return response.choices[0].message.content.strip()
            
    except Exception as e:
        logger.error(f"Error using OpenAI API: {e}")
        # Fallback to simple truncation
        sentences = text.split('. ')
        if len(sentences) > 3:
            return '. '.join(sentences[:3]) + '.'
        return text[:500] + "..." if len(text) > 500 else text

def create_transcript_pdf(transcription: str, summary: str, video_title: str, video_url: str) -> bytes:
    """Create PDF with transcription and summary"""
    if not EXPORT_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF export not available.")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Title
    story.append(Paragraph("Video Transcription Report", title_style))
    story.append(Spacer(1, 12))
    
    # Video info
    story.append(Paragraph(f"<b>Video Title:</b> {video_title}", styles['Normal']))
    story.append(Paragraph(f"<b>Video URL:</b> {video_url}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary section
    story.append(Paragraph("Summary", heading_style))
    story.append(Paragraph(summary, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Full transcript section
    story.append(Paragraph("Full Transcription", heading_style))
    story.append(Paragraph(transcription, styles['Normal']))
    
    doc.build(story)
    output.seek(0)
    return output.getvalue()

def parse_duration(duration_str: str) -> float:
    """Parse YouTube duration from ISO 8601 format (PT1H2M3S) to seconds"""
    import re
    
    # Remove PT prefix
    duration_str = duration_str.replace('PT', '')
    
    # Extract hours, minutes, seconds using regex
    hours = re.search(r'(\d+)H', duration_str)
    minutes = re.search(r'(\d+)M', duration_str)
    seconds = re.search(r'(\d+)S', duration_str)
    
    total_seconds = 0
    if hours:
        total_seconds += int(hours.group(1)) * 3600
    if minutes:
        total_seconds += int(minutes.group(1)) * 60
    if seconds:
        total_seconds += int(seconds.group(1))
    
    return float(total_seconds)

def get_youtube_api_service():
    """Initialize YouTube API service"""
    api_key = os.getenv("YOUTUBE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="YouTube API key not configured")
    
    try:
        return build("youtube", "v3", developerKey=api_key)
    except Exception as e:
        logger.error(f"Failed to initialize YouTube API: {e}")
        raise HTTPException(status_code=500, detail="Failed to initialize YouTube API")

def search_videos_with_api(keyword: str) -> List[Video]:
    """Search videos using YouTube Data API v3"""
    try:
        youtube = get_youtube_api_service()
        
        # Calculate date for last 2 years
        two_years_ago = datetime.now() - timedelta(days=730)
        published_after = two_years_ago.isoformat() + "Z"
        
        # Search for videos
        search_response = youtube.search().list(
            q=keyword,
            part="id,snippet",
            maxResults=50,
            type="video",
            order="viewCount",
            publishedAfter=published_after
        ).execute()
        
        video_ids = [item["id"]["videoId"] for item in search_response.get("items", [])]
        
        if not video_ids:
            return []
        
        # Get detailed video information including duration
        videos_response = youtube.videos().list(
            part="snippet,statistics,contentDetails",
            id=",".join(video_ids)
        ).execute()
        
        videos = []
        for video_item in videos_response.get("items", []):
            snippet = video_item["snippet"]
            statistics = video_item["statistics"]
            content_details = video_item["contentDetails"]
            
            # Parse duration from ISO 8601 format (PT1H2M3S)
            duration_str = content_details.get("duration", "PT0S")
            duration_seconds = parse_duration(duration_str)
            
            # Get comments for this video
            comments = []
            try:
                comments_response = youtube.commentThreads().list(
                    part="snippet",
                    videoId=video_item["id"],
                    maxResults=5,
                    order="relevance"
                ).execute()
                
                for comment_item in comments_response.get("items", []):
                    comment_snippet = comment_item["snippet"]["topLevelComment"]["snippet"]
                    comments.append(Comment(
                        text=comment_snippet["textDisplay"],
                        author=comment_snippet["authorDisplayName"],
                        likes=comment_snippet.get("likeCount", 0)
                    ))
            except Exception as e:
                logger.warning(f"Failed to fetch comments for video {video_item['id']}: {e}")
            
            videos.append(Video(
                title=snippet["title"],
                video_url=f"https://www.youtube.com/watch?v={video_item['id']}",
                views=int(statistics.get("viewCount", 0)),
                likes=int(statistics.get("likeCount", 0)),
                description=snippet["description"],
                comment_count=int(statistics.get("commentCount", 0)),
                top_comments=comments,
                thumbnail_url=snippet["thumbnails"]["high"]["url"],
                duration=duration_seconds
            ))
        
        return videos
        
    except Exception as e:
        logger.error(f"YouTube API error: {e}")
        raise HTTPException(status_code=500, detail=f"YouTube API error: {str(e)}")

def search_videos_with_ytdlp(keyword: str) -> List[Video]:
    """Search videos using yt-dlp as fallback"""
    try:
        # Use yt-dlp to search YouTube
        cmd = [
            "yt-dlp",
            "--dump-json",
            "--no-playlist",
            "--max-downloads", "50",
            "--playlist-items", "1-50",
            f"ytsearch50:{keyword}"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            logger.error(f"yt-dlp error: {result.stderr}")
            raise Exception(f"yt-dlp failed: {result.stderr}")
        
        # Parse JSON output
        videos_data = []
        for line in result.stdout.strip().split('\n'):
            if line.strip():
                try:
                    videos_data.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
        
        videos = []
        for video_data in videos_data[:50]:  # Limit to 50 videos
            # Extract basic info
            title = video_data.get("title", "")
            video_url = video_data.get("webpage_url", "")
            views = video_data.get("view_count", 0)
            likes = video_data.get("like_count", 0)
            description = video_data.get("description", "")
            comment_count = video_data.get("comment_count", 0)
            thumbnail_url = video_data.get("thumbnail", "")
            duration = video_data.get("duration", 0)  # Duration in seconds
            
            # For yt-dlp, we can't easily get comments without additional requests
            # So we'll create empty comments
            comments = []
            
            videos.append(Video(
                title=title,
                video_url=video_url,
                views=views,
                likes=likes,
                description=description,
                comment_count=comment_count,
                top_comments=comments,
                thumbnail_url=thumbnail_url,
                duration=float(duration) if duration else None
            ))
        
        return videos
        
    except Exception as e:
        logger.error(f"yt-dlp error: {e}")
        raise HTTPException(status_code=500, detail=f"yt-dlp error: {str(e)}")

@app.post("/get_videos", response_model=VideoResponse)
async def get_videos(request: VideoRequest):
    """Get top 50 YouTube videos for a keyword"""
    try:
        videos = []
        source = "api"
        
        # Try YouTube API first
        try:
            videos = search_videos_with_api(request.keyword)
            logger.info(f"Successfully fetched {len(videos)} videos using YouTube API")
        except HTTPException as api_error:
            logger.warning(f"YouTube API failed, trying yt-dlp: {api_error}")
            # Fallback to yt-dlp
            try:
                videos = search_videos_with_ytdlp(request.keyword)
                source = "yt-dlp"
                logger.info(f"Successfully fetched {len(videos)} videos using yt-dlp")
            except HTTPException as ytdlp_error:
                logger.error(f"Both API and yt-dlp failed: {ytdlp_error}")
                raise ytdlp_error
        
        return VideoResponse(
            videos=videos,
            total_count=len(videos),
            source=source,
            keyword=request.keyword
        )
        
    except Exception as e:
        logger.error(f"Error in get_videos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/excel")
async def export_to_excel(request: ExportRequest):
    """Export videos to Excel file"""
    try:
        if not request.videos:
            raise HTTPException(status_code=400, detail="No videos to export")
        
        # Use provided keyword or default
        keyword = request.keyword or "youtube_results"
        
        excel_data = create_excel_file(request.videos, keyword)
        filename = generate_filename("yt_results", "xlsx", keyword)
        
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export to Excel: {str(e)}")

@app.post("/export/pdf")
async def export_to_pdf(request: ExportRequest):
    """Export videos to PDF file"""
    try:
        if not request.videos:
            raise HTTPException(status_code=400, detail="No videos to export")
        
        # Use provided keyword or default
        keyword = request.keyword or "youtube_results"
        
        pdf_data = create_pdf_file(request.videos, keyword)
        filename = generate_filename("yt_results", "pdf", keyword)
        
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export to PDF: {str(e)}")

@app.post("/transcribe/{video_url:path}", response_model=TranscribeResponse)
async def transcribe_video(video_url: str):
    """Transcribe a YouTube video"""
    try:
        if not AI_AVAILABLE:
            raise HTTPException(status_code=500, detail="AI libraries not available. Install whisper and transformers.")
        
        # Decode URL
        video_url = urllib.parse.unquote(video_url)
        
        # Validate YouTube URL
        if not ("youtube.com" in video_url or "youtu.be" in video_url):
            raise HTTPException(status_code=400, detail="Invalid YouTube URL. Please provide a valid YouTube link.")
        
        logger.info(f"Starting transcription for video: {video_url}")
        
        # Create temporary directory for audio file
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                # Download audio
                logger.info("Downloading audio...")
                audio_file, video_title, duration = download_audio(video_url, temp_dir)
                
                # Transcribe audio
                logger.info("Transcribing audio...")
                transcription = transcribe_audio(audio_file)
                
                # Clean up temporary files
                try:
                    if os.path.exists(audio_file):
                        os.remove(audio_file)
                except:
                    pass
                
                logger.info(f"Transcription completed for: {video_title}")
                
                return TranscribeResponse(
                    transcription=transcription,
                    video_url=video_url,
                    video_title=video_title,
                    duration=duration
                )
                
            except HTTPException:
                raise
            except Exception as e:
                # Provide more specific error messages
                error_msg = str(e).lower()
                
                if "http error 400" in error_msg or "precondition check failed" in error_msg:
                    raise HTTPException(
                        status_code=400, 
                        detail="YouTube is blocking the download request. This video might be restricted, private, or temporarily unavailable."
                    )
                elif "timeout" in error_msg:
                    raise HTTPException(
                        status_code=408, 
                        detail="Download timeout. The video might be too long or your connection is slow."
                    )
                elif "not found" in error_msg:
                    raise HTTPException(
                        status_code=404, 
                        detail="Video not found. Please check if the YouTube URL is correct and the video is publicly available."
                    )
                elif "audio file is empty" in error_msg:
                    raise HTTPException(
                        status_code=400, 
                        detail="Downloaded audio file is empty. This video might not have audio or might be corrupted."
                    )
                elif "signature extraction failed" in error_msg:
                    raise HTTPException(
                        status_code=400, 
                        detail="YouTube has changed its format. Please try updating yt-dlp or try a different video."
                    )
                else:
                    raise HTTPException(
                        status_code=500, 
                        detail=f"Transcription failed: {str(e)}"
                    )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error transcribing video: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to transcribe video: {str(e)}")

@app.post("/summarize_transcription", response_model=SummarizeResponse)
async def summarize_transcription(request: SummarizeRequest):
    """Summarize a transcription"""
    try:
        if not AI_AVAILABLE:
            raise HTTPException(status_code=500, detail="AI libraries not available. Install transformers.")
        
        if not request.transcription.strip():
            raise HTTPException(status_code=400, detail="Transcription text is empty")
        
        logger.info("Starting text summarization...")
        
        # Summarize the transcription
        summary = summarize_text(request.transcription)
        
        logger.info("Summarization completed")
        
        return SummarizeResponse(
            summary=summary,
            original_length=len(request.transcription),
            summary_length=len(summary)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error summarizing transcription: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to summarize transcription: {str(e)}")

@app.post("/learning_mode/{video_url:path}", response_model=LearningModeResponse)
async def learning_mode(video_url: str):
    """Generate learning flashcards for a YouTube video"""
    try:
        if not AI_AVAILABLE:
            raise HTTPException(status_code=500, detail="AI libraries not available for learning mode.")
        
        # Decode URL
        video_url = urllib.parse.unquote(video_url)
        
        # Validate YouTube URL
        if not ("youtube.com" in video_url or "youtu.be" in video_url):
            raise HTTPException(status_code=400, detail="Invalid YouTube URL. Please provide a valid YouTube link.")
        
        # Extract video ID
        try:
            video_id = extract_video_id(video_url)
        except ValueError as e:
            raise HTTPException(status_code=400, detail="Could not extract video ID from URL. Please check the YouTube URL format.")
        
        logger.info(f"Generating learning mode for video: {video_id}")
        
        # Create temporary directory for audio file
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                # Download audio and get transcription
                logger.info("Downloading audio for learning mode...")
                audio_file, video_title, duration = download_audio(video_url, temp_dir)
                
                # Transcribe audio
                logger.info("Transcribing audio for learning mode...")
                transcription = transcribe_audio(audio_file)
                
                # Check if transcription is long enough for meaningful flashcards
                if len(transcription.strip()) < 100:
                    raise HTTPException(
                        status_code=400, 
                        detail="Transcription is too short to generate meaningful flashcards. The video might be too brief or mostly silent."
                    )
                
                # Generate flashcards
                logger.info("Generating flashcards...")
                flashcards = generate_flashcards_with_gemini(transcription, video_title)
                
                if not flashcards:
                    raise HTTPException(
                        status_code=500, 
                        detail="Failed to generate flashcards. The content might not be suitable for creating learning materials."
                    )
                
                # Clean up temporary files
                try:
                    if os.path.exists(audio_file):
                        os.remove(audio_file)
                except:
                    pass
                
                logger.info(f"Learning mode completed for: {video_title}. Generated {len(flashcards)} flashcards")
                
                return LearningModeResponse(
                    video_id=video_id,
                    video_title=video_title,
                    flashcards=flashcards,
                    total_cards=len(flashcards)
                )
                
            except HTTPException:
                raise
            except Exception as e:
                # Provide more specific error messages
                error_msg = str(e).lower()
                
                if "http error 400" in error_msg or "precondition check failed" in error_msg:
                    raise HTTPException(
                        status_code=400, 
                        detail="YouTube is blocking the download request. This video might be restricted, private, or temporarily unavailable."
                    )
                elif "timeout" in error_msg:
                    raise HTTPException(
                        status_code=408, 
                        detail="Processing timeout. The video might be too long."
                    )
                elif "not found" in error_msg:
                    raise HTTPException(
                        status_code=404, 
                        detail="Video not found. Please check if the YouTube URL is correct and the video is publicly available."
                    )
                else:
                    raise HTTPException(
                        status_code=500, 
                        detail=f"Learning mode failed: {str(e)}"
                    )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in learning mode: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate learning mode: {str(e)}")

@app.post("/export/transcript")
async def export_transcript(
    transcription: str,
    summary: str,
    video_title: str = "Unknown Video",
    video_url: str = "",
    format: str = "pdf"
):
    """Export transcription and summary as PDF or TXT"""
    try:
        if format.lower() == "pdf":
            # Create PDF
            pdf_data = create_transcript_pdf(transcription, summary, video_title, video_url)
            filename = generate_filename("transcript", "pdf", video_title)
            
            return Response(
                content=pdf_data,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        elif format.lower() == "txt":
            # Create TXT content
            txt_content = f"Video Transcription Report\n"
            txt_content += f"{'=' * 50}\n\n"
            txt_content += f"Video Title: {video_title}\n"
            txt_content += f"Video URL: {video_url}\n\n"
            txt_content += f"SUMMARY\n"
            txt_content += f"{'-' * 20}\n"
            txt_content += f"{summary}\n\n"
            txt_content += f"FULL TRANSCRIPTION\n"
            txt_content += f"{'-' * 20}\n"
            txt_content += f"{transcription}\n"
            
            filename = generate_filename("transcript", "txt", video_title)
            
            return Response(
                content=txt_content.encode('utf-8'),
                media_type="text/plain",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'pdf' or 'txt'")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting transcript: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export transcript: {str(e)}")

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    ai_status = "available" if AI_AVAILABLE else "unavailable"
    export_status = "available" if EXPORT_AVAILABLE else "unavailable"
    
    return {
        "status": "healthy", 
        "timestamp": datetime.now().isoformat(),
        "ai_features": ai_status,
        "export_features": export_status
    }

# Add ffmpeg to PATH if it's not found
def ensure_ffmpeg_available():
    """Ensure ffmpeg is available in the PATH"""
    try:
        subprocess.run(["ffmpeg", "-version"], capture_output=True, check=True, timeout=5)
        logger.info("ffmpeg is available in PATH")
    except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
        logger.warning("ffmpeg not found in PATH, trying to add local ffmpeg...")
        
        # Try to find local ffmpeg
        possible_paths = [
            os.path.join(os.path.dirname(__file__), "..", "ffmpeg", "ffmpeg-master-latest-win64-gpl", "bin"),
            os.path.join(os.getcwd(), "ffmpeg", "ffmpeg-master-latest-win64-gpl", "bin"),
            "C:\\Program Files\\ffmpeg\\bin",
            "C:\\ffmpeg\\bin"
        ]
        
        for path in possible_paths:
            ffmpeg_exe = os.path.join(path, "ffmpeg.exe")
            if os.path.exists(ffmpeg_exe):
                logger.info(f"Found ffmpeg at: {path}")
                # Add to PATH
                current_path = os.environ.get("PATH", "")
                if path not in current_path:
                    os.environ["PATH"] = f"{path};{current_path}"
                    logger.info("Added ffmpeg to PATH")
                return
        
        logger.error("Could not find ffmpeg. Transcription may fail.")

# Call this during initialization
ensure_ffmpeg_available()

# Syllabus parsing functions
def parse_text_syllabus(text: str) -> List[SyllabusTopic]:
    """Parse syllabus from plain text"""
    topics = []
    lines = text.strip().split('\n')
    current_unit = "Unit 1"
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Check for unit headers (Unit 1, Chapter 1, etc.)
        unit_match = re.match(r'^(Unit|Chapter|Module|Section)\s*(\d+)[:.\s]*(.*)', line, re.IGNORECASE)
        if unit_match:
            current_unit = f"{unit_match.group(1)} {unit_match.group(2)}"
            topic = unit_match.group(3).strip()
            if topic:
                topics.append(SyllabusTopic(unit=current_unit, topic=topic))
        else:
            # Check for numbered topics (1. Topic, 1) Topic, etc.)
            topic_match = re.match(r'^\d+[.)]\s*(.+)', line)
            if topic_match:
                topic = topic_match.group(1).strip()
                topics.append(SyllabusTopic(unit=current_unit, topic=topic))
            else:
                # Check for bullet points or dashes
                bullet_match = re.match(r'^[-•*]\s*(.+)', line)
                if bullet_match:
                    topic = bullet_match.group(1).strip()
                    topics.append(SyllabusTopic(unit=current_unit, topic=topic))
                else:
                    # Assume it's a topic if it's not empty and doesn't match other patterns
                    if len(line) > 3 and not line.startswith('#'):
                        topics.append(SyllabusTopic(unit=current_unit, topic=line))
    
    return topics

def parse_pdf_syllabus(file_content: bytes) -> List[SyllabusTopic]:
    """Parse syllabus from PDF file"""
    if not SYLLABUS_PARSING_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF parsing not available. Install PyPDF2.")
    
    try:
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        
        return parse_text_syllabus(text)
    except Exception as e:
        logger.error(f"Error parsing PDF: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {str(e)}")

def parse_docx_syllabus(file_content: bytes) -> List[SyllabusTopic]:
    """Parse syllabus from DOCX file"""
    if not SYLLABUS_PARSING_AVAILABLE:
        raise HTTPException(status_code=500, detail="DOCX parsing not available. Install python-docx.")
    
    try:
        docx_file = io.BytesIO(file_content)
        doc = docx.Document(docx_file)
        
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        
        return parse_text_syllabus(text)
    except Exception as e:
        logger.error(f"Error parsing DOCX: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse DOCX: {str(e)}")

def generate_quiz_questions(topics: List[str], num_questions: int = 20, 
                          difficulty: str = "medium", question_types: List[str] = None) -> List[QuizQuestion]:
    """Generate quiz questions using AI with difficulty and type support"""
    if not AI_AVAILABLE:
        raise HTTPException(status_code=500, detail="AI libraries not available for quiz generation.")
    
    try:
        # Use Gemini API for quiz generation
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if gemini_api_key:
            return generate_quiz_with_gemini(topics, num_questions, difficulty, question_types)
        else:
            # Fallback to OpenAI
            openai_api_key = os.getenv("OPENAI_API_KEY")
            if openai_api_key:
                return generate_quiz_with_openai(topics, num_questions, difficulty, question_types)
            else:
                return generate_fallback_quiz(topics, num_questions, difficulty, question_types)
    except Exception as e:
        logger.error(f"Error generating quiz: {e}")
        return generate_fallback_quiz(topics, num_questions, difficulty, question_types)

def generate_quiz_with_gemini(topics: List[str], num_questions: int, 
                            difficulty: str = "medium", question_types: List[str] = None) -> List[QuizQuestion]:
    """Generate quiz using Gemini API with difficulty and type support"""
    try:
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if not gemini_api_key:
            raise Exception("Gemini API key not found")
        
        genai.configure(api_key=gemini_api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        
        topics_text = ", ".join(topics)
        question_types_text = ", ".join(question_types) if question_types else "mcq, true_false"
        
        # Adjust prompt based on difficulty
        difficulty_instruction = {
            "easy": "Create straightforward questions that test basic understanding and recall.",
            "medium": "Create questions that require some analysis and application of concepts.",
            "hard": "Create challenging questions that require deep understanding, synthesis, and critical thinking."
        }.get(difficulty, "Create questions that require some analysis and application of concepts.")
        
        prompt = f"""
As an expert educator and content creator, generate {num_questions} {difficulty}-level quiz questions covering the following topics: {topics_text}

Question types to include: {question_types_text}

{difficulty_instruction}

For each question, ensure:
1. **Clarity and Conciseness**: Questions are easy to understand but require appropriate level of thinking.
2. **Plausible Distractors (for MCQs)**: Options should be relevant and challenging for the {difficulty} level.
3. **Informative Explanations**: The explanation should provide deeper understanding.
4. **Diverse Question Types**: Include the specified question types.
5. **Relevance**: Directly relate to the core concepts within the specified topics.

Return ONLY a valid JSON array with the exact structure below. Include difficulty and type for each question.

[
  {{
    "question": "What is the primary goal of supervised learning in machine learning?",
    "options": ["To group similar data points together without predefined categories", "To find optimal strategies through trial and error interactions", "To learn a mapping from input features to an output target based on labeled examples", "To reduce the dimensionality of data while preserving its structure"],
    "answer": "To learn a mapping from input features to an output target based on labeled examples",
    "explanation": "Supervised learning algorithms are trained on a labeled dataset, meaning each input example is paired with its correct output. The goal is to learn a function that maps inputs to outputs to make predictions on new, unseen data.",
    "topic": "Supervised Learning",
    "difficulty": "{difficulty}",
    "type": "mcq"
  }},
  {{
    "question": "Artificial Intelligence (AI) exclusively focuses on mimicking human intelligence.",
    "options": ["True", "False"],
    "answer": "False",
    "explanation": "While mimicking human intelligence is a significant aspect, AI also encompasses areas like narrow AI (designed for specific tasks, e.g., chess), general AI (human-level intelligence), and superintelligence, which extends beyond human capabilities.",
    "topic": "Introduction to Artificial Intelligence",
    "difficulty": "{difficulty}",
    "type": "true_false"
  }}
]

Do not include any introductory or concluding remarks outside the JSON array.
"""
        
        response = model.generate_content(prompt)
        response_text = response.text.strip()
        
        # Extract JSON from response
        import re
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group()
            questions_data = json.loads(json_str)
            
            questions = []
            for q_data in questions_data:
                if isinstance(q_data, dict) and 'question' in q_data and 'answer' in q_data:
                    questions.append(QuizQuestion(
                        question=q_data.get('question', ''),
                        options=q_data.get('options', []),
                        answer=q_data.get('answer', ''),
                        explanation=q_data.get('explanation', ''),
                        topic=q_data.get('topic', topics[0] if topics else 'General'),
                        difficulty=q_data.get('difficulty', difficulty),
                        type=q_data.get('type', 'mcq')
                    ))
            return questions
        
        return generate_fallback_quiz(topics, num_questions, difficulty, question_types)
        
    except Exception as e:
        logger.error(f"Error generating quiz with Gemini: {e}")
        return generate_fallback_quiz(topics, num_questions, difficulty, question_types)

def generate_quiz_with_openai(topics: List[str], num_questions: int, 
                            difficulty: str = "medium", question_types: List[str] = None) -> List[QuizQuestion]:
    """Generate quiz using OpenAI API with difficulty and type support"""
    try:
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise Exception("OpenAI API key not found")
        
        client = openai.OpenAI(api_key=openai_api_key)
        
        topics_text = ", ".join(topics)
        question_types_text = ", ".join(question_types) if question_types else "mcq, true_false"
        
        prompt = f"""
Create {num_questions} {difficulty}-level quiz questions based on these topics: {topics_text}

Question types to include: {question_types_text}

Return ONLY a valid JSON array with this exact structure:
[
  {{
    "question": "What is the main purpose of machine learning?",
    "options": ["Data storage", "Pattern recognition", "Web browsing", "File compression"],
    "answer": "Pattern recognition",
    "explanation": "Machine learning focuses on identifying patterns in data to make predictions.",
    "topic": "Machine Learning Basics",
    "difficulty": "{difficulty}",
    "type": "mcq"
  }},
  {{
    "question": "Neural networks can only process numerical data.",
    "options": ["True", "False"],
    "answer": "False",
    "explanation": "Neural networks can process various data types including text, images, and audio.",
    "topic": "Neural Networks",
    "difficulty": "{difficulty}",
    "type": "true_false"
  }}
]

Distribute questions evenly across the topics. Make questions appropriate for {difficulty} level.
"""
        
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that creates educational quiz questions."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=2000,
            temperature=0.7
        )
        
        response_text = response.choices[0].message.content.strip()
        
        # Extract JSON from response
        import re
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group()
            questions_data = json.loads(json_str)
            
            questions = []
            for q_data in questions_data:
                if isinstance(q_data, dict) and 'question' in q_data and 'answer' in q_data:
                    questions.append(QuizQuestion(
                        question=q_data.get('question', ''),
                        options=q_data.get('options', []),
                        answer=q_data.get('answer', ''),
                        explanation=q_data.get('explanation', ''),
                        topic=q_data.get('topic', topics[0] if topics else 'General'),
                        difficulty=q_data.get('difficulty', difficulty),
                        type=q_data.get('type', 'mcq')
                    ))
            return questions
        
        return generate_fallback_quiz(topics, num_questions, difficulty, question_types)
        
    except Exception as e:
        logger.error(f"Error generating quiz with OpenAI: {e}")
        return generate_fallback_quiz(topics, num_questions, difficulty, question_types)

def generate_fallback_quiz(topics: List[str], num_questions: int, 
                          difficulty: str = "medium", question_types: List[str] = None) -> List[QuizQuestion]:
    """Generate basic quiz questions as fallback with difficulty and type support"""
    questions = []
    question_types = question_types or ["mcq", "true_false"]
    
    # Question templates for variety
    mcq_templates = [
        ("What is the primary purpose of {topic}?", ["Data analysis", "Problem solving", "Information processing", "All of the above"], "All of the above"),
        ("Which of the following best describes {topic}?", ["A simple concept", "A complex system", "A fundamental principle", "An optional feature"], "A fundamental principle"),
        ("{topic} is most commonly used for:", ["Basic operations", "Advanced applications", "Educational purposes", "Research and development"], "Advanced applications"),
        ("The main advantage of {topic} is:", ["Speed", "Accuracy", "Flexibility", "All of the above"], "All of the above"),
        ("{topic} typically involves:", ["Single step processes", "Multi-step procedures", "Random actions", "No specific method"], "Multi-step procedures")
    ]
    
    true_false_templates = [
        ("{topic} is essential for modern computing systems.", "True"),
        ("{topic} can only be applied in specific scenarios.", "False"),
        ("Understanding {topic} requires advanced mathematical knowledge.", "False"),
        ("{topic} has evolved significantly over the years.", "True"),
        ("{topic} is only relevant for technical professionals.", "False"),
        ("{topic} provides a foundation for other related concepts.", "True"),
        ("{topic} is a static field that doesn't change.", "False"),
        ("{topic} can be learned through practical experience.", "True")
    ]
    
    # Generate questions ensuring good coverage
    questions_per_topic = max(1, num_questions // len(topics)) if topics else num_questions
    
    for i, topic in enumerate(topics):
        topic_questions = 0
        
        # Add MCQ questions for this topic
        if "mcq" in question_types:
            for j, (template, options, answer) in enumerate(mcq_templates):
                if topic_questions >= questions_per_topic:
                    break
                questions.append(QuizQuestion(
                    question=template.format(topic=topic),
                    options=options,
                    answer=answer,
                    explanation=f"{topic} is a comprehensive concept that encompasses multiple aspects including {', '.join(options[:-1])}.",
                    topic=topic,
                    difficulty=difficulty,
                    type="mcq"
                ))
                topic_questions += 1
        
        # Add True/False questions for this topic
        if "true_false" in question_types:
            for j, (template, answer) in enumerate(true_false_templates):
                if topic_questions >= questions_per_topic:
                    break
                questions.append(QuizQuestion(
                    question=template.format(topic=topic),
                    options=["True", "False"],
                    answer=answer,
                    explanation=f"This statement about {topic} is {answer.lower()} because {topic} has specific characteristics and applications.",
                    topic=topic,
                    difficulty=difficulty,
                    type="true_false"
                ))
                topic_questions += 1
    
    # If we need more questions, add general questions
    while len(questions) < num_questions:
        remaining = num_questions - len(questions)
        if remaining >= 2:
            # Add one MCQ and one True/False
            topic = topics[len(questions) % len(topics)] if topics else "General Computing"
            
            # MCQ
            if "mcq" in question_types:
                questions.append(QuizQuestion(
                    question=f"What aspect of {topic} is most important for beginners?",
                    options=["Advanced techniques", "Basic understanding", "Complex algorithms", "Specialized tools"],
                    answer="Basic understanding",
                    explanation=f"Starting with basic understanding of {topic} provides a solid foundation for advanced learning.",
                    topic=topic,
                    difficulty=difficulty,
                    type="mcq"
                ))
            
            # True/False
            if "true_false" in question_types:
                questions.append(QuizQuestion(
                    question=f"{topic} requires continuous learning and adaptation.",
                    options=["True", "False"],
                    answer="True",
                    explanation=f"{topic} is a dynamic field that evolves with technology and requires ongoing education.",
                    topic=topic,
                    difficulty=difficulty,
                    type="true_false"
                ))
        else:
            # Add one final question
            topic = topics[0] if topics else "General Computing"
            question_type = question_types[0] if question_types else "mcq"
            
            if question_type == "mcq":
                questions.append(QuizQuestion(
                    question=f"Which statement about {topic} is correct?",
                    options=["It's only for experts", "It's accessible to everyone", "It's outdated", "It's too simple"],
                    answer="It's accessible to everyone",
                    explanation=f"{topic} is designed to be accessible to learners at all levels with proper guidance.",
                    topic=topic,
                    difficulty=difficulty,
                    type="mcq"
                ))
            else:
                questions.append(QuizQuestion(
                    question=f"{topic} is a fundamental concept in computing.",
                    options=["True", "False"],
                    answer="True",
                    explanation=f"{topic} provides essential knowledge for understanding modern computing systems.",
                    topic=topic,
                    difficulty=difficulty,
                    type="true_false"
                ))
    
    return questions[:num_questions]  # Ensure we don't exceed the requested number

def generate_learning_report(quiz_attempts: List[QuizAttempt], watched_videos: List[str], syllabus_topics: List[SyllabusTopic]) -> ReportResponse:
    """Generate comprehensive learning report"""
    try:
        # Calculate overall score
        total_questions = len(quiz_attempts)
        correct_answers = sum(1 for attempt in quiz_attempts if attempt.is_correct)
        overall_score = (correct_answers / total_questions * 100) if total_questions > 0 else 0
        
        # Calculate topic scores
        topic_scores = {}
        topic_attempts = {}
        for attempt in quiz_attempts:
            if attempt.topic not in topic_attempts:
                topic_attempts[attempt.topic] = []
            topic_attempts[attempt.topic].append(attempt)
        
        for topic, attempts in topic_attempts.items():
            correct = sum(1 for a in attempts if a.is_correct)
            topic_scores[topic] = (correct / len(attempts) * 100) if attempts else 0
        
        # Identify weak areas (topics with score < 70%)
        weak_areas = [topic for topic, score in topic_scores.items() if score < 70]
        
        # Find common mistakes
        mistakes = [attempt for attempt in quiz_attempts if not attempt.is_correct]
        common_mistakes = []
        if mistakes:
            # Group by topic and find most common wrong answers
            mistake_groups = {}
            for mistake in mistakes:
                if mistake.topic not in mistake_groups:
                    mistake_groups[mistake.topic] = []
                mistake_groups[mistake.topic].append(mistake.selected_answer)
            
            for topic, wrong_answers in mistake_groups.items():
                if len(wrong_answers) >= 2:  # At least 2 mistakes in same topic
                    common_mistakes.append(f"Multiple mistakes in {topic}: {', '.join(set(wrong_answers))}")
        
        # Find unwatched topics
        watched_topics = set()
        for video_url in watched_videos:
            # Extract topic from video URL or title (simplified)
            for topic in syllabus_topics:
                if topic.topic.lower() in video_url.lower():
                    watched_topics.add(topic.topic)
        
        unwatched_topics = [topic.topic for topic in syllabus_topics if topic.topic not in watched_topics]
        
        # Generate recommendations
        recommendations = []
        if weak_areas:
            recommendations.append(f"Focus on reviewing: {', '.join(weak_areas[:3])}")
        if unwatched_topics:
            recommendations.append(f"Watch videos on: {', '.join(unwatched_topics[:3])}")
        if overall_score < 80:
            recommendations.append("Consider retaking the quiz after reviewing weak areas")
        if not recommendations:
            recommendations.append("Great job! Keep up the good work.")
        
        # Prepare report data
        report_data = {
            "quiz_summary": {
                "total_questions": total_questions,
                "correct_answers": correct_answers,
                "incorrect_answers": total_questions - correct_answers,
                "accuracy_percentage": round(overall_score, 2)
            },
            "topic_breakdown": topic_scores,
            "performance_analysis": {
                "strongest_topic": max(topic_scores.items(), key=lambda x: x[1])[0] if topic_scores else "N/A",
                "weakest_topic": min(topic_scores.items(), key=lambda x: x[1])[0] if topic_scores else "N/A",
                "topics_covered": len(topic_scores),
                "topics_missed": len([t for t in syllabus_topics if t.topic not in topic_scores])
            },
            "study_recommendations": recommendations,
            "timestamp": datetime.now().isoformat()
        }
        
        return ReportResponse(
            overall_score=overall_score,
            topic_scores=topic_scores,
            weak_areas=weak_areas,
            recommendations=recommendations,
            common_mistakes=common_mistakes,
            unwatched_topics=unwatched_topics,
            report_data=report_data
        )
        
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")

# Syllabus endpoints
@app.post("/upload_syllabus")
async def upload_syllabus(
    file: Optional[UploadFile] = File(None),
    text_content: Optional[str] = Form(None)
):
    """Upload and parse syllabus from file or text"""
    try:
        topics = []
        
        if file:
            # Parse uploaded file
            file_content = await file.read()
            file_extension = file.filename.lower().split('.')[-1] if file.filename else ''
            
            if file_extension == 'pdf':
                topics = parse_pdf_syllabus(file_content)
            elif file_extension in ['docx', 'doc']:
                topics = parse_docx_syllabus(file_content)
            else:
                raise HTTPException(status_code=400, detail="Unsupported file format. Use PDF or DOCX.")
        elif text_content:
            # Parse text content
            topics = parse_text_syllabus(text_content)
        else:
            raise HTTPException(status_code=400, detail="Either file or text_content must be provided.")
        
        if not topics:
            raise HTTPException(status_code=400, detail="No topics found in the syllabus. Please check the format.")
        
        logger.info(f"Successfully parsed {len(topics)} topics from syllabus")
        
        return {
            "topics": topics,
            "total_topics": len(topics),
            "units": list(set(topic.unit for topic in topics))
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading syllabus: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload syllabus: {str(e)}")

@app.post("/videos_by_syllabus", response_model=SyllabusVideosResponse)
async def get_videos_by_syllabus(request: SyllabusUploadRequest):
    """Get videos for each syllabus topic"""
    try:
        syllabus_mapping = []
        total_videos = 0
        
        for topic in request.topics:
            try:
                # Search for videos using the topic name
                search_keyword = f"{topic.topic} {topic.unit}"
                videos = []
                
                # Try YouTube API first
                try:
                    videos = search_videos_with_api(search_keyword)
                    videos = videos[:5]  # Limit to top 5 videos
                except:
                    # Fallback to yt-dlp
                    try:
                        videos = search_videos_with_ytdlp(search_keyword)
                        videos = videos[:5]  # Limit to top 5 videos
                    except:
                        logger.warning(f"Failed to find videos for topic: {topic.topic}")
                        continue
                
                syllabus_mapping.append(SyllabusVideoMapping(
                    topic=topic.topic,
                    unit=topic.unit,
                    videos=videos
                ))
                
                total_videos += len(videos)
                
            except Exception as e:
                logger.warning(f"Error processing topic {topic.topic}: {e}")
                continue
        
        logger.info(f"Found {total_videos} videos for {len(syllabus_mapping)} topics")
        
        return SyllabusVideosResponse(
            syllabus_mapping=syllabus_mapping,
            total_topics=len(syllabus_mapping),
            total_videos=total_videos
        )
        
    except Exception as e:
        logger.error(f"Error getting videos by syllabus: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get videos by syllabus: {str(e)}")

@app.post("/generate_quiz", response_model=QuizResponse)
async def generate_quiz(request: QuizRequest):
    """Generate quiz questions based on topics with offline support"""
    try:
        if not request.topics:
            raise HTTPException(status_code=400, detail="No topics provided for quiz generation.")
        
        # Try to generate quiz using AI
        try:
            questions = generate_quiz_questions(request.topics, request.num_questions, request.difficulty, request.question_types)
            
            if questions:
                # Save quiz to local storage
                try:
                    for topic in request.topics:
                        topic_questions = [q for q in questions if q.topic == topic]
                        if topic_questions:
                            save_quiz_to_storage(
                                request.subject, 
                                "Unit 1",  # Default unit, can be enhanced
                                topic, 
                                topic_questions,
                                request.difficulty,
                                request.question_types
                            )
                except Exception as save_error:
                    logger.warning(f"Failed to save quiz to storage: {save_error}")
                
                # Get unique topics covered
                topics_covered = list(set(q.topic for q in questions))
                
                logger.info(f"Generated {len(questions)} quiz questions covering {len(topics_covered)} topics")
                
                return QuizResponse(
                    questions=questions,
                    total_questions=len(questions),
                    topics_covered=topics_covered,
                    source="api",
                    subject=request.subject,
                    difficulty=request.difficulty,
                    question_types=request.question_types
                )
        
        except Exception as ai_error:
            logger.warning(f"AI quiz generation failed: {ai_error}")
            
            # Try to load from offline storage
            offline_questions = []
            for topic in request.topics:
                offline_quiz = load_quiz_from_storage(request.subject, "Unit 1", topic)
                if offline_quiz:
                    for q_data in offline_quiz.get("questions", []):
                        offline_questions.append(QuizQuestion(
                            question=q_data["question"],
                            options=q_data["options"],
                            answer=q_data["answer"],
                            explanation=q_data.get("explanation"),
                            topic=q_data["topic"],
                            difficulty=q_data.get("difficulty", "medium"),
                            type=q_data.get("type", "mcq")
                        ))
            
            if offline_questions:
                topics_covered = list(set(q.topic for q in offline_questions))
                logger.info(f"Loaded {len(offline_questions)} questions from offline storage")
                
                return QuizResponse(
                    questions=offline_questions,
                    total_questions=len(offline_questions),
                    topics_covered=topics_covered,
                    source="offline",
                    subject=request.subject,
                    difficulty=request.difficulty,
                    question_types=request.question_types
                )
            
            # Fallback to basic questions
            fallback_questions = generate_fallback_quiz(request.topics, request.num_questions)
            topics_covered = list(set(q.topic for q in fallback_questions))
            
            logger.info(f"Using fallback quiz with {len(fallback_questions)} questions")
            
            return QuizResponse(
                questions=fallback_questions,
                total_questions=len(fallback_questions),
                topics_covered=topics_covered,
                source="fallback",
                subject=request.subject,
                difficulty=request.difficulty,
                question_types=request.question_types
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating quiz: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate quiz: {str(e)}")

@app.post("/generate_report", response_model=ReportResponse)
async def generate_report(request: ReportRequest):
    """Generate comprehensive learning report"""
    try:
        if not request.quiz_attempts:
            raise HTTPException(status_code=400, detail="No quiz attempts provided.")
        
        # Generate the report
        report = generate_learning_report(
            request.quiz_attempts,
            request.watched_videos,
            request.syllabus_topics
        )
        
        logger.info(f"Generated learning report with {report.overall_score:.1f}% overall score")
        
        return report
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")

# Create storage directories
def ensure_storage_directories():
    """Create necessary storage directories for offline functionality"""
    directories = [
        "storage/quizzes",
        "storage/materials", 
        "storage/notes",
        "storage/reports",
        "storage/report_templates"
    ]
    
    for directory in directories:
        Path(directory).mkdir(parents=True, exist_ok=True)
    
    # Initialize SQLite database for metadata
    init_storage_db()

def init_storage_db():
    """Initialize SQLite database for storage metadata"""
    db_path = "storage/storage.db"
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create tables
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS quiz_metadata (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject TEXT NOT NULL,
            unit TEXT NOT NULL,
            topic TEXT NOT NULL,
            filename TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            question_count INTEGER,
            difficulty TEXT DEFAULT 'medium',
            question_types TEXT,
            hash TEXT UNIQUE
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS study_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject TEXT NOT NULL,
            topic TEXT NOT NULL,
            material_type TEXT NOT NULL,
            title TEXT NOT NULL,
            url TEXT,
            filename TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            metadata TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            subject TEXT NOT NULL,
            filename TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            score REAL,
            topics_covered TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

def save_quiz_to_storage(subject: str, unit: str, topic: str, questions: List[QuizQuestion], 
                        difficulty: str = "medium", question_types: List[str] = None) -> str:
    """Save quiz to local storage"""
    try:
        # Create subject directory
        subject_dir = Path(f"storage/quizzes/{subject}")
        subject_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate filename
        safe_unit = re.sub(r'[^\w\s-]', '', unit).replace(' ', '_')
        safe_topic = re.sub(r'[^\w\s-]', '', topic).replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_unit}_{safe_topic}_{timestamp}.json"
        filepath = subject_dir / filename
        
        # Prepare quiz data
        quiz_data = {
            "subject": subject,
            "unit": unit,
            "topic": topic,
            "created_at": datetime.now().isoformat(),
            "difficulty": difficulty,
            "question_types": question_types or ["mcq", "true_false"],
            "question_count": len(questions),
            "questions": [
                {
                    "question": q.question,
                    "options": q.options,
                    "answer": q.answer,
                    "explanation": q.explanation,
                    "topic": q.topic,
                    "difficulty": getattr(q, 'difficulty', 'medium'),
                    "type": getattr(q, 'type', 'mcq')
                }
                for q in questions
            ]
        }
        
        # Save to file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(quiz_data, f, indent=2, ensure_ascii=False)
        
        # Save metadata to database
        save_quiz_metadata(subject, unit, topic, filename, len(questions), difficulty, question_types)
        
        return str(filepath)
        
    except Exception as e:
        logger.error(f"Error saving quiz to storage: {e}")
        raise

def save_quiz_metadata(subject: str, unit: str, topic: str, filename: str, 
                      question_count: int, difficulty: str, question_types: List[str]):
    """Save quiz metadata to SQLite database"""
    try:
        conn = sqlite3.connect("storage/storage.db")
        cursor = conn.cursor()
        
        # Generate hash for uniqueness
        content_hash = hashlib.md5(f"{subject}{unit}{topic}{filename}".encode()).hexdigest()
        
        cursor.execute('''
            INSERT OR REPLACE INTO quiz_metadata 
            (subject, unit, topic, filename, question_count, difficulty, question_types, hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (subject, unit, topic, filename, question_count, difficulty, 
              json.dumps(question_types) if question_types else None, content_hash))
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        logger.error(f"Error saving quiz metadata: {e}")

def load_quiz_from_storage(subject: str, unit: str, topic: str) -> Optional[dict]:
    """Load quiz from local storage"""
    try:
        # Find the most recent quiz for this topic
        conn = sqlite3.connect("storage/storage.db")
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT filename FROM quiz_metadata 
            WHERE subject = ? AND unit = ? AND topic = ?
            ORDER BY created_at DESC LIMIT 1
        ''', (subject, unit, topic))
        
        result = cursor.fetchone()
        conn.close()
        
        if not result:
            return None
        
        filename = result[0]
        filepath = Path(f"storage/quizzes/{subject}/{filename}")
        
        if not filepath.exists():
            return None
        
        with open(filepath, 'r', encoding='utf-8') as f:
            quiz_data = json.load(f)
        
        return quiz_data
        
    except Exception as e:
        logger.error(f"Error loading quiz from storage: {e}")
        return None

def get_available_quizzes(subject: str = None) -> List[dict]:
    """Get list of available quizzes from storage"""
    try:
        conn = sqlite3.connect("storage/storage.db")
        cursor = conn.cursor()
        
        if subject:
            cursor.execute('''
                SELECT subject, unit, topic, created_at, question_count, difficulty, question_types
                FROM quiz_metadata 
                WHERE subject = ?
                ORDER BY created_at DESC
            ''', (subject,))
        else:
            cursor.execute('''
                SELECT subject, unit, topic, created_at, question_count, difficulty, question_types
                FROM quiz_metadata 
                ORDER BY created_at DESC
            ''')
        
        results = cursor.fetchall()
        conn.close()
        
        quizzes = []
        for row in results:
            quizzes.append({
                "subject": row[0],
                "unit": row[1], 
                "topic": row[2],
                "created_at": row[3],
                "question_count": row[4],
                "difficulty": row[5],
                "question_types": json.loads(row[6]) if row[6] else []
            })
        
        return quizzes
        
    except Exception as e:
        logger.error(f"Error getting available quizzes: {e}")
        return []

# New endpoints for offline functionality
@app.get("/available_quizzes")
async def get_available_quizzes_endpoint(subject: str = None):
    """Get list of available quizzes from storage"""
    try:
        quizzes = get_available_quizzes(subject)
        return {
            "quizzes": quizzes,
            "total_count": len(quizzes)
        }
    except Exception as e:
        logger.error(f"Error getting available quizzes: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get available quizzes: {str(e)}")

@app.get("/load_quiz/{subject}/{unit}/{topic}")
async def load_quiz_endpoint(subject: str, unit: str, topic: str):
    """Load a specific quiz from storage"""
    try:
        quiz_data = load_quiz_from_storage(subject, unit, topic)
        if not quiz_data:
            raise HTTPException(status_code=404, detail="Quiz not found in storage")
        
        return quiz_data
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading quiz: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load quiz: {str(e)}")

@app.post("/save_study_material")
async def save_study_material(
    subject: str = Form(...),
    topic: str = Form(...),
    material_type: str = Form(...),  # "video", "note", "pdf"
    title: str = Form(...),
    url: str = Form(None),
    file: UploadFile = File(None)
):
    """Save study material to local storage"""
    try:
        # Create directory structure
        material_dir = Path(f"storage/materials/{subject}/{topic}")
        material_dir.mkdir(parents=True, exist_ok=True)
        
        filename = None
        if file:
            # Save uploaded file
            safe_filename = re.sub(r'[^\w\s-]', '', file.filename).replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{timestamp}_{safe_filename}"
            filepath = material_dir / filename
            
            with open(filepath, "wb") as f:
                content = await file.read()
                f.write(content)
        
        # Save metadata to database
        conn = sqlite3.connect("storage/storage.db")
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO study_materials 
            (subject, topic, material_type, title, url, filename, metadata)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (subject, topic, material_type, title, url, filename, json.dumps({})))
        
        conn.commit()
        conn.close()
        
        return {
            "message": "Study material saved successfully",
            "subject": subject,
            "topic": topic,
            "title": title,
            "filename": filename
        }
        
    except Exception as e:
        logger.error(f"Error saving study material: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save study material: {str(e)}")

@app.get("/get_study_materials/{subject}/{topic}")
async def get_study_materials_endpoint(subject: str, topic: str):
    """Get study materials for a specific topic"""
    try:
        conn = sqlite3.connect("storage/storage.db")
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT material_type, title, url, filename, created_at, metadata
            FROM study_materials 
            WHERE subject = ? AND topic = ?
            ORDER BY created_at DESC
        ''', (subject, topic))
        
        results = cursor.fetchall()
        conn.close()
        
        materials = []
        for row in results:
            materials.append({
                "material_type": row[0],
                "title": row[1],
                "url": row[2],
                "filename": row[3],
                "created_at": row[4],
                "metadata": json.loads(row[5]) if row[5] else {}
            })
        
        return {
            "subject": subject,
            "topic": topic,
            "materials": materials,
            "total_count": len(materials)
        }
        
    except Exception as e:
        logger.error(f"Error getting study materials: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get study materials: {str(e)}")

# Initialize storage on startup
ensure_storage_directories()

def run_server(port: int):
    """Run the FastAPI server on a specific port"""
    import uvicorn
    print(f"🚀 Starting FastAPI server on port {port}...")
    uvicorn.run(app, host="0.0.0.0", port=port)

if __name__ == "__main__":
    import threading
    import time
    
    print("🚀 Starting Stu-dih Backend Servers...")
    print("📖 API Documentation: http://localhost:8000/docs or http://localhost:8001/docs")
    print("🔗 Health Check: http://localhost:8000/health or http://localhost:8001/health")
    print("⏹️  Press Ctrl+C to stop")
    
    # Start server on port 8000
    server_8000 = threading.Thread(target=run_server, args=(8000,))
    server_8000.daemon = True
    server_8000.start()
    
    # Start server on port 8001
    server_8001 = threading.Thread(target=run_server, args=(8001,))
    server_8001.daemon = True
    server_8001.start()
    
    try:
        # Keep the main thread alive
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n🛑 Shutting down servers...") 